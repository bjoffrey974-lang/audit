"""
Exports : PDF (reportlab), Excel (openpyxl), SVG, PNG (cairosvg si dispo).
"""
from io import BytesIO
from datetime import datetime

from models import EQUIP_VISUAL, Liaison


# ============================================================================
# SVG  (vue logique ou physique)
# ============================================================================

def generate_svg_string(audit, vue="logique"):
    """Génère le SVG de topologie. Renvoie une chaîne."""
    eqs = sorted(audit.equipements, key=lambda e: e.id)
    liaisons = audit.liaisons

    # Bornes du dessin
    pad = 80
    if eqs:
        max_x = max((e.pos_x or 0) for e in eqs) + 200
        max_y = max((e.pos_y or 0) for e in eqs) + 150
    else:
        max_x, max_y = 1000, 600
    width = max(1000, int(max_x + pad))
    height = max(600, int(max_y + pad))

    parts = []
    parts.append(
        f'<svg xmlns="http://www.w3.org/2000/svg" '
        f'viewBox="0 0 {width} {height}" width="{width}" height="{height}" '
        f'style="background:#fafafa;font-family:Arial,sans-serif">'
    )

    # Cartouche
    parts.append(
        f'<g><rect x="20" y="20" width="380" height="60" rx="6" '
        f'fill="#fff" stroke="#dee2e6"/>'
        f'<text x="35" y="45" font-size="14" font-weight="bold">'
        f'{_xml_escape(audit.client.nom or "Client")}</text>'
        f'<text x="35" y="65" font-size="11" fill="#6c757d">'
        f'Audit du {audit.date_audit} — vue {vue}</text></g>'
    )

    # Liaisons (dessinées d'abord pour être sous les nodes)
    eq_by_id = {e.id: e for e in eqs}
    for l in liaisons:
        s = eq_by_id.get(l.source_id)
        d = eq_by_id.get(l.dest_id)
        if not s or not d:
            continue
        x1, y1 = (s.pos_x or 0) + 60, (s.pos_y or 0) + 40
        x2, y2 = (d.pos_x or 0) + 60, (d.pos_y or 0) + 40
        stroke = "#0d6efd"
        dash = ""
        if l.type == "wifi":
            stroke, dash = "#20c997", 'stroke-dasharray="6,4"'
        elif l.type == "fibre":
            stroke = "#ffc107"
        elif l.type in ("vpn", "sdwan"):
            stroke, dash = "#6610f2", 'stroke-dasharray="4,3"'
        elif l.type == "sip-trunk":
            stroke = "#d63384"
        parts.append(
            f'<line x1="{x1}" y1="{y1}" x2="{x2}" y2="{y2}" '
            f'stroke="{stroke}" stroke-width="2" {dash}/>'
        )
        if l.vlan or l.debit:
            label = " ".join(filter(None, [l.vlan, l.debit]))
            mx, my = (x1 + x2) / 2, (y1 + y2) / 2
            parts.append(
                f'<rect x="{mx-30}" y="{my-10}" width="60" height="14" rx="3" '
                f'fill="#fff" stroke="{stroke}" stroke-width="0.5" opacity="0.9"/>'
                f'<text x="{mx}" y="{my}" text-anchor="middle" '
                f'font-size="10" fill="{stroke}">{_xml_escape(label)}</text>'
            )

    # Relations d'hébergement (VM -> hyperviseur)
    if vue == "physique":
        for host in eqs:
            if host.type != "hyperviseur":
                continue
            vms = [e for e in eqs if e.parent_id == host.id]
            if not vms:
                continue
            members = [host] + vms
            min_x = min((m.pos_x or 0) for m in members) - 16
            min_y = min((m.pos_y or 0) for m in members) - 28
            max_x = max((m.pos_x or 0) for m in members) + 120 + 16
            max_y = max((m.pos_y or 0) for m in members) + 80 + 16
            hv_color = EQUIP_VISUAL.get("hyperviseur", {}).get("color", "#0f5132")
            label = (f"\U0001F4E6 {host.nom_hote or 'Hyperviseur'}"
                     f"{(' — ' + host.modele) if host.modele else ''} "
                     f"({len(vms)} VM)")
            parts.append(
                f'<rect x="{min_x}" y="{min_y}" width="{max_x - min_x}" '
                f'height="{max_y - min_y}" rx="12" fill="{hv_color}" '
                f'fill-opacity="0.05" stroke="{hv_color}" stroke-width="1.5" '
                f'stroke-dasharray="6,4"/>'
                f'<rect x="{min_x}" y="{min_y}" '
                f'width="{min(len(label) * 6.2 + 16, max_x - min_x)}" height="18" '
                f'rx="6" fill="{hv_color}"/>'
                f'<text x="{min_x + 8}" y="{min_y + 13}" font-size="10" '
                f'font-weight="bold" fill="white">{_xml_escape(label)}</text>'
            )
    else:
        for e in eqs:
            if e.parent_id and eq_by_id.get(e.parent_id):
                h = eq_by_id[e.parent_id]
                x1, y1 = (e.pos_x or 0) + 60, (e.pos_y or 0) + 40
                x2, y2 = (h.pos_x or 0) + 60, (h.pos_y or 0) + 40
                mx, my = (x1 + x2) / 2, (y1 + y2) / 2
                parts.append(
                    f'<line x1="{x1}" y1="{y1}" x2="{x2}" y2="{y2}" '
                    f'stroke="#adb5bd" stroke-width="1.5" stroke-dasharray="2,3"/>'
                    f'<rect x="{mx-38}" y="{my-9}" width="76" height="13" rx="3" '
                    f'fill="#fff" stroke="#adb5bd" stroke-width="0.5" opacity="0.9"/>'
                    f'<text x="{mx}" y="{my+1}" text-anchor="middle" '
                    f'font-size="8" fill="#6c757d">h&#233;berg&#233; par</text>'
                )

    # Nœuds
    for e in eqs:
        vis = EQUIP_VISUAL.get(e.type, EQUIP_VISUAL["autre"])
        x, y = e.pos_x or 0, e.pos_y or 0
        nom = e.nom_hote or vis["label"]
        ip = e.ip or ""
        parts.append(
            f'<g transform="translate({x},{y})">'
            f'<rect width="120" height="80" rx="8" fill="white" '
            f'stroke="{vis["color"]}" stroke-width="2"/>'
            f'<rect width="120" height="22" rx="8" fill="{vis["color"]}"/>'
            f'<rect y="14" width="120" height="8" fill="{vis["color"]}"/>'
            f'<text x="60" y="16" text-anchor="middle" font-size="10" '
            f'font-weight="bold" fill="white">{_xml_escape(vis["label"].upper())}</text>'
            f'<text x="60" y="42" text-anchor="middle" font-size="11" '
            f'font-weight="bold">{_xml_escape(nom[:18])}</text>'
            f'<text x="60" y="58" text-anchor="middle" font-size="9" '
            f'fill="#6c757d">{_xml_escape(ip[:18])}</text>'
        )
        if vue == "physique" and (e.rack or e.rack_u):
            tag = " ".join(filter(None, [e.rack, e.rack_u]))
            parts.append(
                f'<text x="60" y="72" text-anchor="middle" font-size="9" '
                f'fill="#6c757d">📍 {_xml_escape(tag)}</text>'
            )
        elif e.vlan:
            parts.append(
                f'<text x="60" y="72" text-anchor="middle" font-size="9" '
                f'fill="#6c757d">VLAN {_xml_escape(e.vlan)}</text>'
            )
        parts.append('</g>')

    # Légende
    legend_x = width - 220
    legend_items = []
    types_present = sorted(set(e.type for e in eqs))
    for i, t in enumerate(types_present[:10]):
        vis = EQUIP_VISUAL.get(t, EQUIP_VISUAL["autre"])
        legend_items.append(
            f'<rect x="{legend_x+10}" y="{40+i*22}" width="14" height="14" '
            f'rx="3" fill="{vis["color"]}"/>'
            f'<text x="{legend_x+32}" y="{52+i*22}" font-size="11">'
            f'{_xml_escape(vis["label"])}</text>'
        )
    if legend_items:
        legend_h = 30 + len(types_present[:10]) * 22
        parts.append(
            f'<g><rect x="{legend_x}" y="20" width="200" height="{legend_h}" '
            f'rx="6" fill="white" stroke="#dee2e6"/>'
            f'<text x="{legend_x+10}" y="35" font-size="11" font-weight="bold">'
            f'Légende</text>'
            + "".join(legend_items) + "</g>"
        )

    parts.append("</svg>")
    return "\n".join(parts)


def generate_svg(audit, vue="logique"):
    buf = BytesIO()
    buf.write(generate_svg_string(audit, vue=vue).encode("utf-8"))
    buf.seek(0)
    return buf


def generate_png(audit, vue="logique"):
    """Convertit le SVG en PNG via cairosvg. Renvoie None si indispo."""
    try:
        import cairosvg
    except ImportError:
        return None
    svg = generate_svg_string(audit, vue=vue)
    png_bytes = cairosvg.svg2png(bytestring=svg.encode("utf-8"), scale=2.0)
    buf = BytesIO(png_bytes)
    buf.seek(0)
    return buf


def _xml_escape(s):
    if s is None:
        return ""
    return (str(s).replace("&", "&amp;").replace("<", "&lt;")
            .replace(">", "&gt;").replace('"', "&quot;"))


# ============================================================================
# PDF (reportlab)
# ============================================================================

def generate_pdf(audit):
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import cm
    from reportlab.lib import colors
    from reportlab.platypus import (
        SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle,
        PageBreak, KeepTogether,
    )

    buf = BytesIO()
    doc = SimpleDocTemplate(
        buf, pagesize=A4,
        leftMargin=2*cm, rightMargin=2*cm,
        topMargin=2*cm, bottomMargin=2*cm,
        title=f"Audit IT - {audit.client.nom}",
    )

    styles = getSampleStyleSheet()
    h1 = ParagraphStyle("H1", parent=styles["Heading1"],
                        textColor=colors.HexColor("#0d3a66"),
                        fontSize=18, spaceAfter=12)
    h2 = ParagraphStyle("H2", parent=styles["Heading2"],
                        textColor=colors.HexColor("#0d6efd"),
                        fontSize=13, spaceAfter=8, spaceBefore=14)
    normal = styles["BodyText"]
    small = ParagraphStyle("small", parent=styles["BodyText"], fontSize=9,
                           textColor=colors.HexColor("#6c757d"))

    story = []

    # Page de garde
    story.append(Spacer(1, 4*cm))
    story.append(Paragraph("COMPTE RENDU<br/>D'AUDIT INFORMATIQUE", h1))
    story.append(Spacer(1, 1*cm))
    story.append(Paragraph(f"<b>Client :</b> {audit.client.nom or '-'}", normal))
    if audit.client.groupe:
        story.append(Paragraph(f"<b>Groupe :</b> {audit.client.groupe}", normal))
    if audit.client.adresse:
        story.append(Paragraph(f"<b>Adresse :</b> {audit.client.adresse}", normal))
    story.append(Spacer(1, 0.5*cm))
    story.append(Paragraph(f"<b>Date :</b> {audit.date_audit}", normal))
    if audit.responsable:
        story.append(Paragraph(f"<b>Responsable de l'audit :</b> {audit.responsable}", normal))
    if audit.interlocuteur_commercial:
        story.append(Paragraph(
            f"<b>Interlocuteur commercial :</b> {audit.interlocuteur_commercial}",
            normal))
    story.append(PageBreak())

    # ----- Contact -----
    story.append(Paragraph("1. Contact", h2))
    c = audit.client
    rows = [
        ["Identité", c.nom or "-"],
        ["Groupe", c.groupe or "-"],
        ["Adresse", c.adresse or "-"],
        ["Contact", c.contact_nom or "-"],
        ["Téléphone", c.contact_tel or "-"],
        ["Email", c.contact_email or "-"],
        ["Horaires", c.horaires or "-"],
    ]
    story.append(_make_kv_table(rows))

    # ----- Activité -----
    story.append(Paragraph("2. Type d'activité", h2))
    story.append(Paragraph(c.activite or "<i>Non renseigné</i>", normal))

    # ----- Sites & Connexions Internet -----
    story.append(Paragraph("3. Sites et connexions Internet", h2))
    if not audit.sites:
        story.append(Paragraph("<i>Aucun site renseigné</i>", normal))
    for site in audit.sites:
        story.append(Paragraph(f"<b>📍 {site.nom}</b>", normal))
        site_rows = [
            ["Adresse", site.adresse or "-"],
            ["FAI", site.fai or "-"],
            ["Type connexion", site.type_connexion or "-"],
            ["Débit (down/up)", f"{site.debit_down or '?'} / {site.debit_up or '?'}"],
            ["IP publique", site.ip_publique or "-"],
            ["Backup Internet", site.backup_connexion or "-"],
            ["Plan IP", site.plan_ip or "-"],
            ["Plage DHCP", site.plage_dhcp or "-"],
        ]
        story.append(_make_kv_table(site_rows))
        story.append(Spacer(1, 0.3*cm))

    # ----- Inventaire équipements -----
    story.append(PageBreak())
    story.append(Paragraph("4. Inventaire des équipements", h2))
    types_groupes = {}
    for e in audit.equipements:
        types_groupes.setdefault(e.type, []).append(e)

    if not audit.equipements:
        story.append(Paragraph("<i>Aucun équipement renseigné</i>", normal))

    for t, items in types_groupes.items():
        label = EQUIP_VISUAL.get(t, {}).get("label", t)
        story.append(Paragraph(f"<b>{label}s ({len(items)})</b>", normal))
        data = [["Nom", "Marque/Modèle", "IP", "VLAN", "Rôle"]]
        for e in items:
            data.append([
                e.nom_hote or "-",
                f"{e.marque or ''} {e.modele or ''}".strip() or "-",
                e.ip or "-",
                e.vlan or "-",
                (e.role or "-")[:50],
            ])
        story.append(_make_data_table(data))
        story.append(Spacer(1, 0.3*cm))

    # ----- Serveurs (détaillé) -----
    serveurs = [e for e in audit.equipements
                if e.type in ("serveur_physique", "serveur_virtuel", "hyperviseur")]
    if serveurs:
        story.append(Paragraph("5. Détail serveurs et virtualisation", h2))
        for s in serveurs:
            story.append(Paragraph(
                f"<b>{s.nom_hote or '?'}</b> ({EQUIP_VISUAL.get(s.type, {}).get('label', s.type)})",
                normal))
            rows = [
                ["OS", f"{s.os or '-'} {s.os_version or ''}"],
                ["IP", s.ip or "-"],
                ["Rôle / fonction", s.role or "-"],
                ["CPU", s.cpu or "-"],
                ["RAM", s.ram or "-"],
                ["Stockage", s.stockage or "-"],
                ["RAID", s.raid or "-"],
                ["Nb VM", str(s.nb_vm) if s.nb_vm else "-"],
            ]
            story.append(_make_kv_table(rows))
            story.append(Spacer(1, 0.2*cm))

    # ----- Stockage / NAS / SAN -----
    storages = [e for e in audit.equipements if e.type in ("nas", "san")]
    if storages:
        story.append(Paragraph("6. Stockage", h2))
        for st in storages:
            rows = [
                ["Type", EQUIP_VISUAL.get(st.type, {}).get("label", st.type)],
                ["Nom", st.nom_hote or "-"],
                ["IP", st.ip or "-"],
                ["Capacité totale", st.capacite_totale or "-"],
                ["Taux occupation", st.taux_occupation or "-"],
                ["RAID", st.raid or "-"],
                ["Date mise en service", st.date_mise_service or "-"],
            ]
            story.append(_make_kv_table(rows))
            story.append(Spacer(1, 0.2*cm))

    # ----- Sauvegardes -----
    story.append(Paragraph("7. Sauvegardes", h2))
    if not audit.sauvegardes:
        story.append(Paragraph("<i>Aucune sauvegarde renseignée</i>", normal))
    else:
        data = [["Périmètre", "Solution", "Type", "Fréquence", "Cible", "Rétention"]]
        for b in audit.sauvegardes:
            data.append([
                b.perimetre or "-", b.solution or "-",
                b.type_sauvegarde or "-", b.frequence or "-",
                b.cible or "-", b.retention or "-",
            ])
        story.append(_make_data_table(data))

    # ----- Sécurité -----
    story.append(PageBreak())
    story.append(Paragraph("8. Sécurité", h2))
    for site in audit.sites:
        story.append(Paragraph(f"<b>📍 {site.nom}</b>", normal))
        rows = [
            ["Antivirus", f"{site.antivirus_solution or '-'} "
                          f"({site.antivirus_gestion or '-'})"],
            ["VPN", f"{site.vpn_type or '-'}"],
            ["Utilisateurs VPN", site.vpn_utilisateurs or "-"],
            ["IDS/IPS", "Oui" if site.idsips_actif else "Non"],
            ["Accès à distance", site.methode_acces_distant or "-"],
        ]
        story.append(_make_kv_table(rows))
        story.append(Spacer(1, 0.3*cm))

    # ----- Téléphonie -----
    voip_sites = [s for s in audit.sites if s.voip_solution or s.voip_pabx]
    ipbx_equip = [e for e in audit.equipements if e.type == "ipbx"]
    if voip_sites or ipbx_equip:
        story.append(Paragraph("9. Infrastructure téléphonique", h2))
        for s in voip_sites:
            rows = [
                ["Site", s.nom],
                ["Solution VoIP", s.voip_solution or "-"],
                ["PABX", s.voip_pabx or "-"],
                ["Softphone", "Oui" if s.voip_softphone else "Non"],
            ]
            story.append(_make_kv_table(rows))
        for e in ipbx_equip:
            rows = [
                ["IPBX", e.nom_hote or "-"],
                ["IP", e.ip or "-"],
                ["Marque/Modèle", f"{e.marque or ''} {e.modele or ''}".strip() or "-"],
            ]
            story.append(_make_kv_table(rows))

    # ----- Imprimantes -----
    imps = [e for e in audit.equipements if e.type in ("imprimante", "scanner")]
    if imps:
        story.append(Paragraph("10. Imprimantes et scanners", h2))
        data = [["Type", "Nom", "Marque", "Modèle", "IP"]]
        for e in imps:
            data.append([
                EQUIP_VISUAL.get(e.type, {}).get("label", e.type),
                e.nom_hote or "-", e.marque or "-", e.modele or "-", e.ip or "-",
            ])
        story.append(_make_data_table(data))

    # ----- Applications métier -----
    story.append(Paragraph("11. Applications métier", h2))
    if not audit.applications:
        story.append(Paragraph("<i>Aucune application renseignée</i>", normal))
    else:
        data = [["Nom", "Type", "Éditeur", "Version", "Support", "Criticité"]]
        for a in audit.applications:
            data.append([
                a.nom or "-", a.type or "-", a.editeur or "-",
                a.version or "-", a.fournisseur_support or "-",
                a.criticite or "-",
            ])
        story.append(_make_data_table(data))

    # ----- Messagerie -----
    story.append(Paragraph("12. Messagerie", h2))
    if not audit.messageries:
        story.append(Paragraph("<i>Aucune messagerie renseignée</i>", normal))
    for m in audit.messageries:
        rows = [
            ["Solution", m.solution or "-"],
            ["Protocole", m.protocole or "-"],
            ["Domaine", m.domaine or "-"],
            ["Nombre de boîtes", str(m.nb_boites) if m.nb_boites else "-"],
            ["Filtre antispam", m.filtre_antispam or "-"],
            ["MX records", m.mx_records or "-"],
        ]
        story.append(_make_kv_table(rows))
        if m.adresses:
            story.append(Paragraph(f"<b>Adresses :</b> {m.adresses}", small))

    # ----- Problèmes en suspens -----
    story.append(Paragraph("13. Problèmes récurrents / points en suspens", h2))
    story.append(Paragraph(
        audit.problemes_recurrents or "<i>Aucun élément renseigné</i>",
        normal))

    # ----- Documentation existante -----
    story.append(Paragraph("14. Documentation existante", h2))
    story.append(Paragraph(
        audit.documentation_existante or "<i>Aucune documentation renseignée</i>",
        normal))

    # ----- Notes générales -----
    if audit.notes_generales:
        story.append(Paragraph("15. Notes générales", h2))
        story.append(Paragraph(audit.notes_generales, normal))

    # ----- Conformité sécurité -----
    if audit.conformites:
        import json as _json
        from conformite_ref import controle_by_id
        story.append(PageBreak())
        story.append(Paragraph("Conformité sécurité des machines", h2))
        story.append(Paragraph(
            "Référentiel maison aligné sur le Guide d'hygiène informatique de l'ANSSI. "
            "Statuts : Conforme / Attention / Critique / Indéterminé.", small))
        for conf in audit.conformites:
            niveau_txt = {
                "conforme": "CONFORME", "partiel": "PARTIELLEMENT CONFORME",
                "non_conforme": "NON CONFORME", "indetermine": "INDÉTERMINÉ",
            }.get(conf.niveau, conf.niveau or "—")
            score_txt = "—" if conf.score is None else f"{conf.score}/100"
            story.append(Spacer(1, 0.3*cm))
            story.append(Paragraph(
                f"<b>{conf.machine or 'Machine'}</b> ({conf.profil or '?'}) — "
                f"score {score_txt} — <b>{niveau_txt}</b> — "
                f"{conf.nb_critiques or 0} point(s) critique(s)", normal))
            try:
                resultats = _json.loads(conf.resultats_json) if conf.resultats_json else []
            except Exception:
                resultats = []
            data = [["Contrôle", "Statut", "Réf. ANSSI", "Détail"]]
            statut_lbl = {"ok": "Conforme", "attention": "Attention",
                          "critique": "Critique", "indetermine": "Indéterminé",
                          "na": "N/A"}
            for r in resultats:
                ctrl = controle_by_id(r.get("id")) or {}
                data.append([
                    ctrl.get("libelle", r.get("id", "?")),
                    statut_lbl.get(r.get("statut"), r.get("statut", "?")),
                    ctrl.get("anssi", "")[:30],
                    (r.get("detail") or "")[:40],
                ])
            story.append(_make_data_table(data))

    # Pied
    story.append(Spacer(1, 1*cm))
    story.append(Paragraph(
        f"<i>Document généré le {datetime.now():%d/%m/%Y à %H:%M}.</i>",
        small))

    doc.build(story)
    buf.seek(0)
    return buf


def _make_kv_table(rows):
    from reportlab.platypus import Table, TableStyle
    from reportlab.lib import colors
    from reportlab.lib.units import cm
    t = Table(rows, colWidths=[4.5*cm, 12*cm])
    t.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (0, -1), colors.HexColor("#f1f3f5")),
        ("TEXTCOLOR", (0, 0), (0, -1), colors.HexColor("#212529")),
        ("FONTNAME", (0, 0), (0, -1), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 9),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#dee2e6")),
        ("LEFTPADDING", (0, 0), (-1, -1), 6),
        ("RIGHTPADDING", (0, 0), (-1, -1), 6),
        ("TOPPADDING", (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
    ]))
    return t


def _make_data_table(data):
    from reportlab.platypus import Table, TableStyle
    from reportlab.lib import colors
    t = Table(data, repeatRows=1)
    t.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#0d6efd")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1),
         [colors.white, colors.HexColor("#f8f9fa")]),
        ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#dee2e6")),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("LEFTPADDING", (0, 0), (-1, -1), 4),
        ("RIGHTPADDING", (0, 0), (-1, -1), 4),
    ]))
    return t


# ============================================================================
# EXCEL (openpyxl)
# ============================================================================

def generate_excel(audit):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()

    # Style entête
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="0D6EFD")
    thin = Side(style="thin", color="DEE2E6")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left = Alignment(horizontal="left", vertical="top", wrap_text=True)

    def _write_sheet(ws, headers, rows):
        for col, h in enumerate(headers, 1):
            c = ws.cell(row=1, column=col, value=h)
            c.font = header_font
            c.fill = header_fill
            c.alignment = center
            c.border = border
        for i, r in enumerate(rows, 2):
            for j, v in enumerate(r, 1):
                cell = ws.cell(row=i, column=j, value=v)
                cell.alignment = left
                cell.border = border
        # Largeur des colonnes
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 22
        ws.freeze_panes = "A2"

    # ----- Feuille Synthèse -----
    ws = wb.active
    ws.title = "Synthèse"
    ws["A1"] = "AUDIT IT — SYNTHÈSE"
    ws["A1"].font = Font(bold=True, size=14, color="0D3A66")
    ws.merge_cells("A1:D1")

    info = [
        ("Client", audit.client.nom),
        ("Groupe", audit.client.groupe),
        ("Adresse", audit.client.adresse),
        ("Contact", audit.client.contact_nom),
        ("Téléphone", audit.client.contact_tel),
        ("Email", audit.client.contact_email),
        ("Date audit", str(audit.date_audit)),
        ("Responsable", audit.responsable),
        ("Activité", audit.client.activite),
    ]
    for i, (k, v) in enumerate(info, 3):
        ws.cell(row=i, column=1, value=k).font = Font(bold=True)
        ws.cell(row=i, column=2, value=v or "-")
    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 50

    # Compteurs
    ws.cell(row=15, column=1, value="STATISTIQUES").font = Font(bold=True, size=12)
    counts = {}
    for e in audit.equipements:
        counts[e.type] = counts.get(e.type, 0) + 1
    row_i = 16
    for t, n in sorted(counts.items()):
        label = EQUIP_VISUAL.get(t, {}).get("label", t)
        ws.cell(row=row_i, column=1, value=label)
        ws.cell(row=row_i, column=2, value=n)
        row_i += 1

    # ----- Feuille Sites -----
    ws = wb.create_sheet("Sites")
    _write_sheet(ws, [
        "Nom", "Adresse", "FAI", "Type connexion", "Débit ↓", "Débit ↑",
        "IP publique", "Backup", "Plan IP", "DHCP",
        "VoIP", "PABX", "Antivirus", "VPN", "IDS/IPS",
    ], [[
        s.nom, s.adresse, s.fai, s.type_connexion, s.debit_down, s.debit_up,
        s.ip_publique, s.backup_connexion, s.plan_ip, s.plage_dhcp,
        s.voip_solution, s.voip_pabx, s.antivirus_solution,
        s.vpn_type, "Oui" if s.idsips_actif else "Non",
    ] for s in audit.sites])

    # ----- Feuille Inventaire équipements -----
    ws = wb.create_sheet("Inventaire")
    _write_sheet(ws, [
        "Type", "Nom hôte", "Marque", "Modèle", "N° série",
        "IP", "MAC", "VLAN", "Rôle",
        "OS", "Version OS", "CPU", "RAM", "Stockage", "RAID",
        "Rack", "U", "Commentaires",
    ], [[
        EQUIP_VISUAL.get(e.type, {}).get("label", e.type),
        e.nom_hote, e.marque, e.modele, e.numero_serie,
        e.ip, e.mac, e.vlan, e.role,
        e.os, e.os_version, e.cpu, e.ram, e.stockage, e.raid,
        e.rack, e.rack_u, e.commentaires,
    ] for e in audit.equipements])

    # ----- Feuille Liaisons -----
    ws = wb.create_sheet("Liaisons")
    eq_by_id = {e.id: e for e in audit.equipements}
    _write_sheet(ws, [
        "Source", "Port src", "Destination", "Port dst",
        "Type", "VLAN", "Débit", "Commentaire",
    ], [[
        (eq_by_id.get(l.source_id).nom_hote
         if eq_by_id.get(l.source_id) else "?"),
        l.port_source,
        (eq_by_id.get(l.dest_id).nom_hote
         if eq_by_id.get(l.dest_id) else "?"),
        l.port_dest,
        l.type, l.vlan, l.debit, l.commentaire,
    ] for l in audit.liaisons])

    # ----- Applications -----
    ws = wb.create_sheet("Applications")
    _write_sheet(ws, [
        "Nom", "Type", "Éditeur", "Version", "Support",
        "Contact support", "Criticité", "Commentaires",
    ], [[
        a.nom, a.type, a.editeur, a.version, a.fournisseur_support,
        a.contact_support, a.criticite, a.commentaires,
    ] for a in audit.applications])

    # ----- Sauvegardes -----
    ws = wb.create_sheet("Sauvegardes")
    _write_sheet(ws, [
        "Périmètre", "Type", "Fréquence", "Solution", "Cible",
        "Rétention", "Dernière vérif", "Chiffré", "Externalisé", "Commentaires",
    ], [[
        b.perimetre, b.type_sauvegarde, b.frequence, b.solution, b.cible,
        b.retention, b.derniere_verification,
        "Oui" if b.chiffrement else "Non",
        "Oui" if b.externalisation else "Non",
        b.commentaires,
    ] for b in audit.sauvegardes])

    # ----- Messagerie -----
    ws = wb.create_sheet("Messagerie")
    _write_sheet(ws, [
        "Solution", "Protocole", "Domaine", "Boîtes",
        "Antispam", "MX records", "Adresses", "Commentaires",
    ], [[
        m.solution, m.protocole, m.domaine, m.nb_boites,
        m.filtre_antispam, m.mx_records, m.adresses, m.commentaires,
    ] for m in audit.messageries])

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf
